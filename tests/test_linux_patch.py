import json
import os

import pytest

import linux_patch as lp
from ssh_exec import Result

FIX = os.path.join(os.path.dirname(__file__), "fixtures")


def fixture(name):
    with open(os.path.join(FIX, name)) as fh:
        return fh.read()


# --- section splitter -------------------------------------------------------

def test_split_sections_ignores_leading_noise():
    sec = lp.split_sections(fixture("discover_web01.txt"))
    assert set(sec) >= {"FACTS", "UPDATES", "SECURITY", "REBOOT", "HISTORY"}
    # the sudo noise before the first fence is dropped
    assert all("noise" not in "\n".join(v) for v in sec.values())


def test_split_sections_end_closes():
    sec = lp.split_sections("===A===\nx\n===END===\ntrailing")
    assert sec == {"A": ["x"]}


# --- facts / updates / security --------------------------------------------

def test_parse_facts():
    sec = lp.split_sections(fixture("discover_web01.txt"))
    facts = lp.parse_facts(sec["FACTS"])
    assert facts["hostname"] == "web01.hostname.loc"
    assert facts["pm"] == "dnf"
    assert "Red Hat" in facts["distro"]


def test_parse_updates_splits_name_and_arch():
    sec = lp.split_sections(fixture("discover_web01.txt"))
    ups = lp.parse_updates(sec["UPDATES"])
    names = {u["name"] for u in ups}
    assert names == {"openssl", "openssl-libs", "kernel", "sudo", "vim-minimal"}
    ossl = next(u for u in ups if u["name"] == "openssl")
    assert ossl["arch"] == "x86_64"
    assert ossl["version"] == "1:3.0.7-25.el9_3"
    assert ossl["repo"] == "rhel-9-baseos"


def test_parse_security_maps_severity_by_name():
    sec = lp.split_sections(fixture("discover_web01.txt"))
    m = lp.parse_security(sec["SECURITY"])
    assert m["openssl"] == ("RHSA-2024:0012", "Important")
    assert m["kernel"][1] == "Critical"
    assert "vim-minimal" not in m  # no advisory -> not security


def test_parse_reboot():
    sec = lp.split_sections(fixture("discover_web01.txt"))
    assert lp.parse_reboot(sec["REBOOT"]) is True
    assert lp.parse_reboot(["reboot_required=0"]) is False
    assert lp.parse_reboot(["reboot_required=unknown"]) is None


def test_parse_history_id():
    sec = lp.split_sections(fixture("discover_web01.txt"))
    assert lp.parse_history_id(sec["HISTORY"]) == 47


# --- host_record ------------------------------------------------------------

def test_host_record_tags_security_and_counts():
    res = Result("web01.hostname.loc", ok=True, stdout=fixture("discover_web01.txt"))
    rec = lp.host_record(res)
    assert rec["reachable"] is True
    assert rec["counts"]["total"] == 5
    assert rec["counts"]["security"] == 4  # openssl, openssl-libs, kernel, sudo
    assert rec["reboot_required"] is True
    # security updates sort first
    assert rec["updates"][0]["security"] is True
    vim = next(u for u in rec["updates"] if u["name"] == "vim-minimal")
    assert vim["security"] is False and vim["advisory"] is None


def test_host_record_unreachable():
    res = Result("dead.host", ok=False, error="Connection timed out")
    rec = lp.host_record(res)
    assert rec["reachable"] is False
    assert rec["error"] == "Connection timed out"
    assert rec["counts"]["total"] == 0


# --- apply reconciliation ---------------------------------------------------

def test_parse_apply_result_reconciles_drift():
    # plan had 5 packages; vim-minimal is NOT in the live PRECHECK set -> drifted
    planned = [{"name": n} for n in
               ["openssl", "openssl-libs", "kernel", "sudo", "vim-minimal"]]
    res = Result("web01.hostname.loc", ok=True, stdout=fixture("apply_web01.txt"))
    out = lp.parse_apply_result(res, planned)
    assert out["status"] == "applied"
    assert set(out["applied"]) == {"openssl", "openssl-libs", "kernel", "sudo"}
    assert out["skipped_current"] == ["vim-minimal"]
    assert out["txn"] == 48
    assert out["install_rc"] == 0


def test_parse_apply_result_failed_rc():
    planned = [{"name": "openssl"}]
    body = ("===PRECHECK===\nopenssl.x86_64\n===INSTALL===\n"
            "Error: nothing provides foo\ninstall_rc=1\n===TXN===\n===END===\n")
    out = lp.parse_apply_result(Result("h", ok=True, stdout=body), planned)
    assert out["status"] == "failed"
    assert out["install_rc"] == 1


def test_parse_apply_result_all_current():
    planned = [{"name": "openssl"}]
    body = ("===PRECHECK===\n===INSTALL===\nnothing to do\ninstall_rc=0\n"
            "===TXN===\n===END===\n")
    out = lp.parse_apply_result(Result("h", ok=True, stdout=body), planned)
    assert out["status"] == "already_current"
    assert out["applied"] == []


def test_parse_apply_result_detects_scheduled_reboot():
    planned = [{"name": "openssl"}]
    body = ("===PRECHECK===\nopenssl.x86_64\n===INSTALL===\ninstall_rc=0\n"
            "===TXN===\n50\n===POSTREBOOT===\nreboot_required=1\n"
            "===REBOOT-SCHEDULED===\n===END===\n")
    out = lp.parse_apply_result(Result("h", ok=True, stdout=body), planned)
    assert out["rebooted"] is True


# --- script builders --------------------------------------------------------

def test_install_script_security_and_reboot_flags():
    s = lp.build_install_script(["openssl", "kernel"], True, True)
    assert "--security update openssl kernel" in s
    assert "systemctl reboot" in s
    s2 = lp.build_install_script(["openssl"], False, False)
    assert "--security" not in s2
    assert "systemctl reboot" not in s2


def test_install_script_quotes_hostile_names():
    s = lp.build_install_script(["a; rm -rf /"], False, False)
    assert "'a; rm -rf /'" in s
    assert "; rm -rf / " not in s.replace("'a; rm -rf /'", "")


def test_rollback_script():
    assert "history undo 42" in lp.build_rollback_script(42)


# --- plan assembly / IO -----------------------------------------------------

def test_build_plan_summary():
    recs = [lp.host_record(Result("web01", ok=True,
                                  stdout=fixture("discover_web01.txt"))),
            lp.host_record(Result("dead", ok=False, error="timeout"))]
    plan = lp.build_plan(recs, security_only=True)
    s = plan["summary"]
    assert s["hosts_total"] == 2
    assert s["hosts_reachable"] == 1
    assert s["hosts_failed"] == 1
    assert s["updates_total"] == 5
    assert s["updates_security"] == 4
    assert s["hosts_reboot_required"] == 1
    assert plan["options"]["security_only"] is True
    assert plan["schema"] == lp.PLAN_SCHEMA


def test_plan_roundtrip(tmp_path):
    recs = [lp.host_record(Result("web01", ok=True,
                                  stdout=fixture("discover_web01.txt")))]
    plan = lp.build_plan(recs)
    p = tmp_path / "plan.json"
    lp.write_plan(str(p), plan)
    loaded = lp.load_plan(str(p))
    assert loaded["summary"]["updates_total"] == 5


def test_load_plan_rejects_wrong_schema(tmp_path):
    p = tmp_path / "bad.json"
    p.write_text(json.dumps({"schema": "something-else"}))
    with pytest.raises(ValueError):
        lp.load_plan(str(p))


# --- Excel report -----------------------------------------------------------

def test_write_report_is_formula_clean(tmp_path):
    from xlsx_safe import verify
    recs = [lp.host_record(Result("web01", ok=True,
                                  stdout=fixture("discover_web01.txt"))),
            lp.host_record(Result("dead", ok=False, error="timeout"))]
    plan = lp.build_plan(recs)
    out = tmp_path / "r.xlsx"
    lp.write_report(str(out), plan)
    assert verify(str(out)) == {}
    from openpyxl import load_workbook
    wb = load_workbook(str(out))
    assert "Summary" in wb.sheetnames
    assert "Security Advisories" in wb.sheetnames
    assert "Reboot Required" in wb.sheetnames


def test_report_neutralizes_formula_injection(tmp_path):
    # a hostile package name that starts with '=' must not become a formula
    res = Result("h", ok=True, stdout=(
        "===FACTS===\nhostname=h\npm=dnf\n===UPDATES===\n"
        "=cmd|'/C calc'.x86_64|1.0|repo\n===SECURITY===\n===REBOOT===\n"
        "reboot_required=0\n===HISTORY===\n1\n===END===\n"))
    plan = lp.build_plan([lp.host_record(res)])
    out = tmp_path / "r.xlsx"
    lp.write_report(str(out), plan)
    from xlsx_safe import verify
    assert verify(str(out)) == {}
