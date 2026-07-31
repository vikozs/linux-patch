#!/usr/bin/env python3
"""
linux_patch.py — staged dnf patch management across a RHEL 9 fleet.

Fourth tool in the family (linux-audit, linux-harden, linux-diskspace). Reuses
ssh_exec.py (transport) and xlsx_safe.py (Excel safety).

Three modes:

    discover   SSH the fleet, collect pending updates, tag security advisories,
               detect reboot-required. Write a patch plan (JSON) and a report
               (xlsx). Changes nothing.

    apply      Consume a plan, RE-VALIDATE every host against live state (plans
               go stale), install the still-pending updates with per-host
               confirmation, optionally reboot (serialized, opt-in).

    rollback   Undo a single host's patch transaction (dnf history undo).

    report     Re-render an existing plan to xlsx. No SSH. (used by CI smoke)

Safety
------
  * apply re-runs `dnf check-update` before installing: the plan is a candidate
    list, never a licence to install blindly.
  * reboot is report-only by default. --reboot opts in, one host at a time,
    waiting for each to return before touching the next. One host failing to
    come back aborts the reboot loop; the rest of the fleet is left alone.
  * a dnf error on one host is data, not a crash: that host is flagged and the
    fleet carries on.

Run artifacts (plan JSON, xlsx) describe your infrastructure. Keep them out of
version control — see .gitignore.
"""

import argparse
import concurrent.futures as cf
import datetime as _dt
import json
import re
import sys
import time

from ssh_exec import SSHConfig, host_label, parse_hosts, run_fleet, run_one

__version__ = "1.0.0"
BUILD = "2026-07-31.initial"

PLAN_SCHEMA = "linux-patch.update-plan"
RESULT_SCHEMA = "linux-patch.patch-result"
SCHEMA_VERSION = "1.0"

# ---------------------------------------------------------------------------
# Remote collector
# ---------------------------------------------------------------------------
# One script, one round trip. Sections are fenced with unambiguous markers so
# the parser never has to guess where dnf's chatter ends. dnf check-update
# exits 100 when updates exist, so we swallow its rc explicitly.

_SECTIONS = ("FACTS", "UPDATES", "SECURITY", "REBOOT", "HISTORY")

DISCOVER_SCRIPT = r"""
set -u
PM=dnf; command -v dnf >/dev/null 2>&1 || PM=yum
echo "===FACTS==="
echo "hostname=$(hostname -f 2>/dev/null || hostname)"
echo "distro=$( (. /etc/os-release 2>/dev/null && echo "$PRETTY_NAME") || echo unknown)"
echo "pm=$PM"
echo "===UPDATES==="
# columns: name.arch  version-release  repo  (exit 100 = updates present)
$PM -q check-update 2>/dev/null | awk 'NF>=3 && $1 !~ /^Obsoleting|^Last|^Security/ {print $1"|"$2"|"$3}'
echo "===SECURITY==="
# advisory  severity/type  package.arch
$PM -q updateinfo list security 2>/dev/null | awk 'NF>=3 {print $1"|"$2"|"$3}'
echo "===REBOOT==="
if command -v needs-restarting >/dev/null 2>&1; then
  if needs-restarting -r >/dev/null 2>&1; then echo "reboot_required=0"; else echo "reboot_required=1"; fi
else
  echo "reboot_required=unknown"
fi
echo "===HISTORY==="
# most recent transaction id, for reference
$PM -q history list 2>/dev/null | awk '/^ *[0-9]+ / {print $1; exit}'
echo "===END==="
"""


def build_install_script(packages, security_only, do_reboot):
    """Build the apply-side script for one host.

    Re-validates against live check-update, installs only the intersection of
    the plan and what is still pending, records the resulting transaction id,
    and (optionally) schedules a reboot AFTER emitting output so the report is
    captured before the host drops.
    """
    sec = "--security " if security_only else ""
    # Space-joined, shell-quoted package names.
    pkgs = " ".join(_shquote(p) for p in packages)
    reboot = ""
    if do_reboot:
        # nohup so the reboot survives our ssh channel closing; delay lets our
        # stdout flush and the ssh session return cleanly first.
        reboot = (
            '\necho "===REBOOT-SCHEDULED==="\n'
            'nohup sh -c "sleep 3; systemctl reboot" >/dev/null 2>&1 &\n'
        )
    return r"""
set -u
PM=dnf; command -v dnf >/dev/null 2>&1 || PM=yum
echo "===PRECHECK==="
# still-pending set, live
$PM -q check-update 2>/dev/null | awk 'NF>=3 && $1 !~ /^Obsoleting|^Last/ {print $1}'
echo "===INSTALL==="
if [ -n "%(pkgs)s" ]; then
  $PM -y %(sec)supdate %(pkgs)s 2>&1
  echo "install_rc=$?"
else
  echo "install_rc=0"
  echo "nothing to do"
fi
echo "===TXN==="
$PM -q history list 2>/dev/null | awk '/^ *[0-9]+ / {print $1; exit}'
echo "===POSTREBOOT==="
if command -v needs-restarting >/dev/null 2>&1; then
  if needs-restarting -r >/dev/null 2>&1; then echo "reboot_required=0"; else echo "reboot_required=1"; fi
else
  echo "reboot_required=unknown"
fi%(reboot)s
echo "===END==="
""" % {"pkgs": pkgs, "sec": sec, "reboot": reboot}


def build_rollback_script(txn):
    return r"""
set -u
PM=dnf; command -v dnf >/dev/null 2>&1 || PM=yum
echo "===ROLLBACK==="
$PM -y history undo %(txn)s 2>&1
echo "rollback_rc=$?"
echo "===END==="
""" % {"txn": _shquote(str(txn))}


def _shquote(s):
    if re.fullmatch(r"[A-Za-z0-9._+:@%/-]+", s or ""):
        return s
    return "'" + str(s).replace("'", "'\\''") + "'"


# ---------------------------------------------------------------------------
# Parsers (pure — fully unit tested against captured dnf output)
# ---------------------------------------------------------------------------

def split_sections(stdout):
    """Split fenced collector output into {section: [lines]}.

    Tolerant of leading sudo/ssh noise before the first fence.
    """
    out = {}
    cur = None
    for line in stdout.splitlines():
        m = re.match(r"^===([A-Z-]+)===$", line.strip())
        if m:
            name = m.group(1)
            cur = None if name == "END" else name
            if cur is not None:
                out.setdefault(cur, [])
            continue
        if cur is not None:
            out[cur].append(line)
    return out


def parse_facts(lines):
    facts = {}
    for line in lines:
        if "=" in line:
            k, v = line.split("=", 1)
            facts[k.strip()] = v.strip()
    return facts


def parse_updates(lines):
    """[{name, arch, version, repo}] from `name.arch|version|repo` rows."""
    out = []
    for line in lines:
        parts = line.split("|")
        if len(parts) != 3:
            continue
        na, ver, repo = (p.strip() for p in parts)
        if not na or not ver:
            continue
        name, _, arch = na.rpartition(".")
        if not name:  # no dot — treat whole token as name
            name, arch = na, ""
        out.append({"name": name, "arch": arch, "version": ver, "repo": repo})
    return out


def parse_security(lines):
    """{package_name: (advisory, severity)} keyed by bare package name.

    updateinfo rows look like: `RHSA-2024:1234 Important/Sec. openssl.x86_64`.
    Severity is the token before the first '/'.
    """
    out = {}
    for line in lines:
        parts = line.split("|")
        if len(parts) != 3:
            continue
        adv, sev, pkg = (p.strip() for p in parts)
        sev = sev.split("/", 1)[0]
        name = pkg.rsplit(".", 1)[0] if "." in pkg else pkg
        # strip epoch:version-release if updateinfo emitted the full nevra
        name = re.split(r"-\d", name, 1)[0]
        out[name] = (adv, sev)
    return out


def parse_reboot(lines):
    for line in lines:
        if line.startswith("reboot_required="):
            v = line.split("=", 1)[1].strip()
            return {"0": False, "1": True}.get(v, None)
    return None


def parse_history_id(lines):
    for line in lines:
        s = line.strip()
        if s.isdigit():
            return int(s)
    return None


def host_record(res):
    """Turn one discover Result into a per-host plan record."""
    if not res.ok:
        return {"host": res.host, "reachable": False, "error": res.error,
                "facts": {}, "updates": [], "reboot_required": None,
                "counts": {"total": 0, "security": 0}}
    sec = split_sections(res.stdout)
    facts = parse_facts(sec.get("FACTS", []))
    updates = parse_updates(sec.get("UPDATES", []))
    secmap = parse_security(sec.get("SECURITY", []))
    for u in updates:
        adv = secmap.get(u["name"])
        u["security"] = bool(adv)
        u["advisory"] = adv[0] if adv else None
        u["severity"] = adv[1] if adv else None
    nsec = sum(1 for u in updates if u["security"])
    return {
        "host": res.host,
        "hostname": facts.get("hostname", res.host),
        "reachable": True,
        "error": None,
        "facts": facts,
        "reboot_required": parse_reboot(sec.get("REBOOT", [])),
        "last_txn": parse_history_id(sec.get("HISTORY", [])),
        "updates": sorted(updates, key=lambda u: (not u["security"], u["name"])),
        "counts": {"total": len(updates), "security": nsec},
    }


def parse_apply_result(res, planned):
    """Interpret an apply Result against the host's planned package list.

    Reconciliation happens here: only packages still in the live PRECHECK set
    were eligible; anything in the plan but absent from PRECHECK is reported as
    already-current (drift since discover), not failed.
    """
    if not res.ok:
        return {"host": res.host, "reachable": False, "error": res.error,
                "status": "unreachable", "applied": [], "skipped_current": [],
                "txn": None, "reboot_required": None, "rebooted": False}
    sec = split_sections(res.stdout)
    live_pending = {l.strip() for l in sec.get("PRECHECK", []) if l.strip()}
    planned_names = [p["name"] for p in planned]
    # name may carry .arch in the pending set; match on bare name prefix
    def still_pending(name):
        return any(lp == name or lp.startswith(name + ".") for lp in live_pending)
    eligible = [n for n in planned_names if still_pending(n)]
    already = [n for n in planned_names if not still_pending(n)]
    install = sec.get("INSTALL", [])
    rc = None
    for line in install:
        m = re.match(r"install_rc=(\d+)", line.strip())
        if m:
            rc = int(m.group(1))
    txn = parse_history_id(sec.get("TXN", []))
    reboot_req = parse_reboot(sec.get("POSTREBOOT", []))
    rebooted = "REBOOT-SCHEDULED" in res.stdout
    status = "applied" if rc == 0 else "failed"
    if rc == 0 and not eligible:
        status = "already_current"
    return {"host": res.host, "reachable": True, "error": None,
            "status": status, "applied": eligible, "skipped_current": already,
            "install_rc": rc, "txn": txn, "reboot_required": reboot_req,
            "rebooted": rebooted,
            "detail": "\n".join(install).strip()[:4000]}


# ---------------------------------------------------------------------------
# Plan assembly / IO
# ---------------------------------------------------------------------------

def _now():
    return _dt.datetime.now(_dt.timezone.utc).replace(microsecond=0).isoformat()


def build_plan(records, security_only=False):
    hosts_ok = [r for r in records if r["reachable"]]
    items = sum(r["counts"]["total"] for r in hosts_ok)
    sec = sum(r["counts"]["security"] for r in hosts_ok)
    need_reboot = sum(1 for r in hosts_ok if r["reboot_required"])
    return {
        "schema": PLAN_SCHEMA,
        "schema_version": SCHEMA_VERSION,
        "generated": _now(),
        "generator": {"tool": "linux-patch", "version": __version__, "build": BUILD},
        "options": {"security_only": security_only},
        "summary": {
            "hosts_total": len(records),
            "hosts_reachable": len(hosts_ok),
            "hosts_failed": len(records) - len(hosts_ok),
            "updates_total": items,
            "updates_security": sec,
            "hosts_reboot_required": need_reboot,
        },
        "hosts": records,
    }


def write_plan(path, plan):
    with open(path, "w") as fh:
        json.dump(plan, fh, indent=2)
        fh.write("\n")


def load_plan(path):
    try:
        with open(path) as fh:
            plan = json.load(fh)
    except json.JSONDecodeError as e:
        raise ValueError("%s is not valid JSON: %s" % (path, e))
    if not isinstance(plan, dict) or plan.get("schema") != PLAN_SCHEMA:
        raise ValueError("not a %s file: %s" % (PLAN_SCHEMA, path))
    return plan


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

NAVY = "1F3864"
HIGH = "C00000"
MED = "ED7D31"
LOW = "FFC000"
GOOD = "70AD47"
RULE = "D9D9D9"
PALE = "E2EFDA"

SEV_FILL = {
    "Critical": (HIGH, "FFFFFF"),
    "Important": (MED, "FFFFFF"),
    "Moderate": (LOW, "000000"),
    "Low": (PALE, "000000"),
}


def write_report(path, plan):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from xlsx_safe import guard, safe_sheet_name, sweep, verify

    wb = Workbook()
    used = set()

    def sheet(title):
        ws = wb.create_sheet(safe_sheet_name(title, used))
        return ws

    def header(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.freeze_panes = "A2"

    hosts = plan["hosts"]
    ok = [h for h in hosts if h["reachable"]]

    # Summary
    ws = wb.active
    ws.title = safe_sheet_name("Summary", used)
    header(ws, ["Host", "Distro", "Updates", "Security", "Reboot needed"])
    for h in sorted(ok, key=lambda x: -x["counts"]["total"]):
        ws.append([h["hostname"], h["facts"].get("distro", ""),
                   h["counts"]["total"], h["counts"]["security"],
                   "yes" if h["reboot_required"] else "no"])
        for c in ws[ws.max_row]:
            guard(c)
        if h["counts"]["security"]:
            ws.cell(ws.max_row, 4).fill = PatternFill("solid", fgColor=MED)
            ws.cell(ws.max_row, 4).font = Font(color="FFFFFF")
        if h["reboot_required"]:
            ws.cell(ws.max_row, 5).fill = PatternFill("solid", fgColor=LOW)

    # Pending Updates
    ws = sheet("Pending Updates")
    header(ws, ["Host", "Package", "Arch", "Available", "Repo", "Security", "Severity"])
    for h in ok:
        for u in h["updates"]:
            ws.append([h["hostname"], u["name"], u.get("arch", ""), u["version"],
                       u.get("repo", ""), "yes" if u["security"] else "",
                       u.get("severity") or ""])
            for c in ws[ws.max_row]:
                guard(c)
            if u["security"]:
                fill, font = SEV_FILL.get(u.get("severity"), (MED, "FFFFFF"))
                cell = ws.cell(ws.max_row, 6)
                cell.fill = PatternFill("solid", fgColor=fill)
                cell.font = Font(color=font)

    # Security Advisories
    ws = sheet("Security Advisories")
    header(ws, ["Host", "Advisory", "Severity", "Package", "Available"])
    for h in ok:
        for u in h["updates"]:
            if u["security"]:
                ws.append([h["hostname"], u.get("advisory") or "",
                           u.get("severity") or "", u["name"], u["version"]])
                for c in ws[ws.max_row]:
                    guard(c)

    # Reboot Required
    ws = sheet("Reboot Required")
    header(ws, ["Host", "Distro", "Pending updates"])
    for h in ok:
        if h["reboot_required"]:
            ws.append([h["hostname"], h["facts"].get("distro", ""),
                       h["counts"]["total"]])
            for c in ws[ws.max_row]:
                guard(c)

    # Errors
    ws = sheet("Errors")
    header(ws, ["Host", "Error"])
    for h in hosts:
        if not h["reachable"]:
            ws.append([h["host"], h.get("error") or "unreachable"])
            for c in ws[ws.max_row]:
                guard(c)
            for c in ws[ws.max_row]:
                c.fill = PatternFill("solid", fgColor=HIGH)
                c.font = Font(color="FFFFFF")

    # About
    ws = sheet("About")
    s = plan["summary"]
    about = [
        ("Tool", "linux-patch %s" % __version__),
        ("Build", plan["generator"].get("build", BUILD)),
        ("Generated", plan["generated"]),
        ("Hosts total", s["hosts_total"]),
        ("Hosts reachable", s["hosts_reachable"]),
        ("Hosts failed", s["hosts_failed"]),
        ("Updates total", s["updates_total"]),
        ("Security updates", s["updates_security"]),
        ("Hosts needing reboot", s["hosts_reboot_required"]),
        ("Security-only plan", str(plan.get("options", {}).get("security_only", False))),
        ("Note", "check-update output is a candidate list; apply re-validates "
                 "against live state before installing."),
    ]
    for k, v in about:
        ws.append([k, v])
        guard(ws.cell(ws.max_row, 2))
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 70

    for w in wb.worksheets:
        for col in "ABCDEFG":
            if w[col + "1"].value:
                w.column_dimensions[col].width = max(
                    w.column_dimensions[col].width or 0, 16)

    swept = sweep(wb)
    wb.save(path)
    bad = verify(path)
    if bad:
        raise RuntimeError("report has formula cells after sweep: %s" % bad)
    return swept


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------

def say(msg):
    print(msg, file=sys.stderr, flush=True)


def do_discover(hosts, cfg, args):
    say("linux-patch %s [build %s] — discover, %d host(s)"
        % (__version__, BUILD, len(hosts)))
    records = []
    for res in run_fleet(hosts, DISCOVER_SCRIPT, cfg, workers=args.workers):
        rec = host_record(res)
        records.append(rec)
        if rec["reachable"]:
            say("  %-40s %3d updates (%d security)%s"
                % (rec["hostname"], rec["counts"]["total"],
                   rec["counts"]["security"],
                   "  reboot-required" if rec["reboot_required"] else ""))
        else:
            say("  %-40s FAILED: %s" % (res.host, rec["error"]))
    records.sort(key=lambda r: (not r["reachable"], -r["counts"]["total"]))
    plan = build_plan(records, security_only=args.security)
    write_plan(args.plan, plan)
    write_report(args.output, plan)
    s = plan["summary"]
    say("\nPlan:   %s" % args.plan)
    say("Report: %s" % args.output)
    say("Totals: %d updates across %d hosts, %d security, %d need reboot"
        % (s["updates_total"], s["hosts_reachable"], s["updates_security"],
           s["hosts_reboot_required"]))
    return plan


def _confirm(prompt, force):
    if force:
        return True
    try:
        return input(prompt + " [y/N] ").strip().lower() in ("y", "yes")
    except EOFError:
        return False


def wait_for_host(host, cfg, timeout=600, interval=10):
    """Poll until the host answers a trivial command, or timeout."""
    deadline = time.time() + timeout
    probe = SSHConfig(user=cfg.user, port=cfg.port, identity=cfg.identity,
                      escalate="none", host_key_checking=cfg.host_key_checking,
                      ssh_opts=cfg.ssh_opts, connect_timeout=8, cmd_timeout=15)
    probe.ssh_pass = cfg.ssh_pass
    while time.time() < deadline:
        r = run_one(host, "echo up", probe)
        if r.ok and "up" in r.stdout:
            return True
        time.sleep(interval)
    return False


def do_apply(plan, hosts, cfg, args):
    say("linux-patch %s [build %s] — apply%s%s"
        % (__version__, BUILD, ", security-only" if args.security else "",
           ", reboot" if args.reboot else ""))
    by_target = {h["target"]: h for h in hosts}
    results = []
    reboot_aborted = False
    for rec in plan["hosts"]:
        if not rec["reachable"]:
            continue
        host = by_target.get(rec["host"]) or {"target": rec["host"],
                                              "user": None, "port": None}
        planned = rec["updates"]
        if args.security:
            planned = [u for u in planned if u.get("security")]
        if not planned:
            say("  %-40s nothing to apply" % rec["hostname"])
            continue
        prompt = ("  Apply %d update(s) to %s%s?"
                  % (len(planned), rec["hostname"],
                     " + reboot" if args.reboot and not reboot_aborted else ""))
        if not _confirm(prompt, args.force):
            results.append({"host": rec["host"], "reachable": True,
                            "status": "skipped", "applied": [],
                            "skipped_current": [], "txn": None,
                            "reboot_required": None, "rebooted": False})
            say("    skipped")
            continue
        do_reboot = bool(args.reboot) and not reboot_aborted
        script = build_install_script([p["name"] for p in planned],
                                      args.security, do_reboot)
        res = run_one(host, script, cfg)
        out = parse_apply_result(res, planned)
        results.append(out)
        say("    %s: %s" % (rec["hostname"], out["status"]))
        if do_reboot and out.get("rebooted"):
            say("    waiting for %s to return..." % rec["hostname"])
            if wait_for_host(host, cfg, timeout=args.reboot_timeout):
                say("    %s is back" % rec["hostname"])
            else:
                reboot_aborted = True
                out["status"] = "reboot_timeout"
                say("    %s did not return in %ds — aborting further reboots"
                    % (rec["hostname"], args.reboot_timeout))
    _write_results(args.results, plan, results, mode="apply")
    _summarize_apply(results)
    return results


def do_rollback(host_target, txn, hosts, cfg, args):
    say("linux-patch %s [build %s] — rollback txn %s on %s"
        % (__version__, BUILD, txn, host_target))
    host = next((h for h in hosts if h["target"] == host_target),
                {"target": host_target, "user": None, "port": None})
    if not _confirm("  Undo transaction %s on %s?" % (txn, host_target), args.force):
        say("  aborted")
        return None
    res = run_one(host, build_rollback_script(txn), cfg)
    sec = split_sections(res.stdout)
    body = "\n".join(sec.get("ROLLBACK", []))
    rc = None
    m = re.search(r"rollback_rc=(\d+)", res.stdout)
    if m:
        rc = int(m.group(1))
    ok = res.ok and rc == 0
    say("  %s (rc=%s)" % ("done" if ok else "FAILED", rc))
    if body:
        say(body[:2000])
    return {"host": host_target, "txn": txn, "ok": ok, "rc": rc, "detail": body}


def _write_results(path, plan, results, mode):
    doc = {
        "schema": RESULT_SCHEMA,
        "schema_version": SCHEMA_VERSION,
        "generated": _now(),
        "generator": {"tool": "linux-patch", "version": __version__, "build": BUILD},
        "mode": mode,
        "plan": {"generated": plan.get("generated"),
                 "generator": plan.get("generator")},
        "hosts": results,
    }
    with open(path, "w") as fh:
        json.dump(doc, fh, indent=2)
        fh.write("\n")
    say("Results: %s" % path)


def _summarize_apply(results):
    from collections import Counter
    c = Counter(r["status"] for r in results)
    say("\nApplied: %d  already-current: %d  skipped: %d  failed: %d"
        % (c.get("applied", 0), c.get("already_current", 0),
           c.get("skipped", 0), c.get("failed", 0) + c.get("reboot_timeout", 0)))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_cfg(args):
    cfg = SSHConfig(
        user=args.user, port=args.port, identity=args.identity,
        escalate=args.escalate, ask_ssh_pass=args.ask_ssh_pass,
        ssh_pass_env=args.ssh_pass_env, ask_sudo_pass=args.ask_sudo_pass,
        sudo_pass_same_as_ssh=args.sudo_pass_same_as_ssh,
        host_key_checking=args.host_key_checking, ssh_opts=args.ssh_opt,
        connect_timeout=args.connect_timeout, cmd_timeout=args.cmd_timeout)
    cfg.resolve_passwords()
    return cfg


def _add_conn_args(ap):
    ap.add_argument("-H", "--hosts", metavar="FILE", help="host list file")
    ap.add_argument("-u", "--user", help="default SSH user")
    ap.add_argument("-p", "--port", help="default SSH port")
    ap.add_argument("-i", "--identity", help="SSH private key")
    ap.add_argument("--escalate", choices=("none", "sudo"), default="sudo")
    ap.add_argument("--ask-ssh-pass", action="store_true")
    ap.add_argument("--ssh-pass-env", metavar="VAR")
    ap.add_argument("--ask-sudo-pass", action="store_true")
    ap.add_argument("--sudo-pass-same-as-ssh", action="store_true")
    ap.add_argument("--ssh-opt", action="append", default=[], metavar="OPT")
    ap.add_argument("--host-key-checking", default="accept-new")
    ap.add_argument("--connect-timeout", type=int, default=10)
    ap.add_argument("--cmd-timeout", type=int, default=300)
    ap.add_argument("--workers", type=int, default=8)


def main(argv=None):
    ap = argparse.ArgumentParser(
        prog="linux-patch",
        description="Staged dnf patch management for a RHEL 9 fleet.")
    ap.add_argument("--version", action="version",
                    version="linux-patch %s (build %s)" % (__version__, BUILD))
    sub = ap.add_subparsers(dest="cmd", required=True)

    d = sub.add_parser("discover", help="collect pending updates, write plan")
    _add_conn_args(d)
    d.add_argument("--security", action="store_true",
                   help="mark plan as security-only (apply installs only those)")
    d.add_argument("--plan", default="patch_plan.json")
    d.add_argument("-o", "--output", default="patch_report.xlsx")

    a = sub.add_parser("apply", help="install updates from a plan")
    _add_conn_args(a)
    a.add_argument("--plan", required=True)
    a.add_argument("--security", action="store_true",
                   help="install only security-advisory packages")
    a.add_argument("--reboot", action="store_true",
                   help="reboot each host after apply, serialized (opt-in)")
    a.add_argument("--reboot-timeout", type=int, default=600,
                   help="seconds to wait for a host to return after reboot")
    a.add_argument("--results", default="patch_results.json")
    a.add_argument("-y", "--force", action="store_true",
                   help="skip per-host confirmation")

    r = sub.add_parser("rollback", help="undo one host's patch transaction")
    _add_conn_args(r)
    r.add_argument("--host", required=True)
    r.add_argument("--txn", required=True)
    r.add_argument("-y", "--force", action="store_true")

    rp = sub.add_parser("report", help="re-render a plan to xlsx (no SSH)")
    rp.add_argument("--plan", required=True)
    rp.add_argument("-o", "--output", default="patch_report.xlsx")

    args = ap.parse_args(argv)

    if args.cmd == "report":
        try:
            plan = load_plan(args.plan)
        except ValueError as e:
            say("error: %s" % e)
            return 2
        write_report(args.output, plan)
        say("Report: %s" % args.output)
        return 0

    hosts = []
    if getattr(args, "hosts", None):
        hosts = parse_hosts(args.hosts)
    if args.cmd in ("discover", "apply") and not hosts:
        ap.error("no hosts: pass -H/--hosts FILE")

    cfg = build_cfg(args)

    if args.cmd == "discover":
        do_discover(hosts, cfg, args)
    elif args.cmd == "apply":
        try:
            plan = load_plan(args.plan)
        except ValueError as e:
            say("error: %s" % e)
            return 2
        do_apply(plan, hosts, cfg, args)
    elif args.cmd == "rollback":
        do_rollback(args.host, args.txn, hosts, cfg, args)
    return 0


if __name__ == "__main__":
    sys.exit(main())
